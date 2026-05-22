#!/usr/bin/env python3
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

def create_xlsx_template(path, title):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data Report"
    
    # Header Row
    headers = ["ID", "Metric Name", "Value", "Currency", "Status"]
    ws.append(headers)
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Sample Data
    data = [
        [1, "Conversion Rate", 3.5, "USD", "On Track"],
        [2, "Customer Acquisition Cost", 120, "USD", "At Risk"],
        [3, "Monthly Recurring Revenue", 15000, "USD", "On Track"],
    ]
    
    for row in data:
        ws.append(row)
        
    # Currency Formatting for Value column (Column C)
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=3).number_format = '"$"#,##0.00'

    # Auto-column width
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    wb.save(path)

if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("Usage: python3 generate_xlsx.py <path>")
        sys.exit(1)
    create_xlsx_template(sys.argv[1], "Reporting Template")
