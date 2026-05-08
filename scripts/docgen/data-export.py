#!/usr/bin/env python3
"""
data-export.py — AInchors Data Export (XLSX)
Usage: python3 data-export.py --title "Title" --output /path/out.xlsx [--data /path/data.json]
"""
import argparse
import json
import sys
from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
SUBHEADER_FILL = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
SUBHEADER_FONT = Font(bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
CENTER = Alignment(horizontal='center', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)


def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 60)


def style_header_row(ws, row_num, num_cols):
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--title', required=True)
    parser.add_argument('--output', required=True)
    parser.add_argument('--data', default=None)
    args = parser.parse_args()

    data = {}
    if args.data:
        with open(args.data) as f:
            data = json.load(f)

    title = data.get('title', args.title)
    doc_date = date.today().strftime('%d %B %Y')
    metrics = data.get('metrics', {
        'Total Projects': 12,
        'Active Clients': 7,
        'Recommendations Delivered': 34,
        'Avg ROI Reported': '320%',
        'Report Date': doc_date,
    })
    headers = data.get('headers', ['Client', 'Project', 'Status', 'Value (AUD)', 'Completion %'])
    rows = data.get('rows', [
        ['Acme Corp', 'AI Strategy', 'In Progress', '$45,000', '65%'],
        ['Beta Industries', 'Process Automation', 'Completed', '$38,000', '100%'],
        ['Gamma Ltd', 'Data Platform', 'Planning', '$72,000', '10%'],
        ['Delta Solutions', 'ML Pipeline', 'In Progress', '$55,000', '40%'],
    ])

    wb = openpyxl.Workbook()

    # ── SUMMARY SHEET ─────────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = 'Summary'

    # Title banner
    ws_sum.merge_cells('A1:B1')
    title_cell = ws_sum['A1']
    title_cell.value = f'{title} — {doc_date}'
    title_cell.font = Font(bold=True, size=14, color='FFFFFF')
    title_cell.fill = PatternFill(start_color='0D1117', end_color='0D1117', fill_type='solid')
    title_cell.alignment = CENTER
    ws_sum.row_dimensions[1].height = 28

    # Sub-header
    ws_sum['A2'] = 'Metric'
    ws_sum['B2'] = 'Value'
    style_header_row(ws_sum, 2, 2)
    ws_sum.row_dimensions[2].height = 20

    # Metrics rows
    for i, (key, val) in enumerate(metrics.items(), start=3):
        ws_sum[f'A{i}'] = key
        ws_sum[f'B{i}'] = val
        ws_sum[f'A{i}'].border = THIN_BORDER
        ws_sum[f'B{i}'].border = THIN_BORDER
        ws_sum[f'A{i}'].alignment = LEFT
        ws_sum[f'B{i}'].alignment = LEFT
        # Alternating row fill
        if i % 2 == 0:
            ws_sum[f'A{i}'].fill = SUBHEADER_FILL
            ws_sum[f'B{i}'].fill = SUBHEADER_FILL

    auto_width(ws_sum)

    # ── DATA SHEET ────────────────────────────────────────────────────────────
    ws_data = wb.create_sheet('Data')

    # Title banner
    ws_data.merge_cells(f'A1:{get_column_letter(len(headers))}1')
    title_cell_d = ws_data['A1']
    title_cell_d.value = f'{title} — Raw Data'
    title_cell_d.font = Font(bold=True, size=13, color='FFFFFF')
    title_cell_d.fill = PatternFill(start_color='0D1117', end_color='0D1117', fill_type='solid')
    title_cell_d.alignment = CENTER
    ws_data.row_dimensions[1].height = 24

    # Header row
    for c, hdr in enumerate(headers, 1):
        cell = ws_data.cell(row=2, column=c, value=hdr)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    ws_data.row_dimensions[2].height = 20

    # Data rows
    for r_idx, row in enumerate(rows, start=3):
        for c_idx, val in enumerate(row, start=1):
            cell = ws_data.cell(row=r_idx, column=c_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = LEFT
            if r_idx % 2 == 0:
                cell.fill = SUBHEADER_FILL

    auto_width(ws_data)

    wb.save(args.output)
    print(f'Data export saved: {args.output}')


if __name__ == '__main__':
    main()
